#!/usr/bin/env python3
"""
Equity Dashboard Generator

Reads MK Stock Analysis.xlsx (SaaS sheet) and generates:
1. Obsidian Markdown table (SaaS Dashboard.md)
2. CSV backup (SaaS Data.csv)

Calculates: Rule of 40, Rule of 40 (FCF), Dilution-Adj FCF Margin, 
P/S to Rof40, P/S to Growth, MSS Score, MSS Rating
"""

import os
import re
import csv
from datetime import datetime

# Configuration
EXCEL_FILE = "/home/art/Obsidian/04-Reference/Finance/Equity Research/MK Stock Analysis.xlsx"
OUTPUT_DIR = "/home/art/Obsidian/04-Reference/Finance/Equity Research"
MARKDOWN_FILE = os.path.join(OUTPUT_DIR, "SaaS Dashboard.md")
CSV_FILE = os.path.join(OUTPUT_DIR, "SaaS Data.csv")

def normalize_quarter(q_str):
    """Normalize quarter format to Q[1-4]/YY."""
    if not q_str: return q_str
    q_str = str(q_str).strip()
    match = re.match(r'Q(\d)/(\d{2,4})', q_str)
    if match:
        quarter = match.group(1)
        year_str = match.group(2)
        year = int(year_str)
        # Convert 4-digit year to 2-digit
        if year >= 2000:
            year = year - 2000
        elif year >= 1900:
            year = year - 1900
        return f"Q{quarter}/{year:02d}"
    return q_str

def get_quarter_order(q_str):
    """Convert quarter string (e.g., 'Q3/2025' or 'Q1/22') to sortable tuple (year, quarter)."""
    if not q_str: return (9999, 9)
    q_str = str(q_str).strip()
    # Match Q[1-4]/[YY or YYYY]
    match = re.match(r'Q(\d)/(\d{2,4})', q_str)
    if match:
        quarter = int(match.group(1))
        year_str = match.group(2)
        year = int(year_str)
        # Convert 2-digit year to 4-digit (assume 2000s for 00-69, 1900s for 70-99)
        if year < 100:
            if year < 70:
                year += 2000
            else:
                year += 1900
        return (year, quarter)
    return (9999, 9)

def clean_value(val):
    """Clean cell value for display."""
    if val is None:
        return "-"
    if isinstance(val, float):
        return round(val, 2)
    return str(val)

def format_percent(val):
    """Format as percentage string."""
    if val is None:
        return "-"
    if isinstance(val, float):
        if abs(val) <= 1.0 and val > -1.0:
            return f"{val*100:.1f}%"
        return f"{val:.1f}%"
    return str(val)

def format_currency(val):
    """Format as currency string."""
    if val is None:
        return "-"
    if isinstance(val, (int, float)):
        return f"${val:.2f}"
    return str(val)

# --- Calculation Functions ---

def calculate_rule_of_40(growth, ops_margin):
    """Rule of 40 = Growth + Ops Margin"""
    if growth is None or ops_margin is None:
        return None
    return growth + ops_margin

def calculate_rule_of_40_fcf(growth, fcf_margin):
    """Rule of 40 (FCF) = Growth + FCF Margin"""
    if growth is None or fcf_margin is None:
        return None
    return growth + fcf_margin

def calculate_dilution_adj_fcf(fcf_margin, sbc):
    """Dilution-Adj FCF Margin = FCF Margin - SBC %"""
    if fcf_margin is None or sbc is None:
        return None
    return fcf_margin - sbc

def calculate_ps_to_rof40(ps, rule40):
    """P/S to Rof40 = P/S / Rule of 40"""
    if ps is None or rule40 is None or rule40 == 0:
        return None
    return ps / rule40

def calculate_ps_to_growth(ps, growth):
    """P/S to Growth = P/S / Growth"""
    if ps is None or growth is None or growth == 0:
        return None
    return ps / growth

def calculate_mss(growth, nrr, rule40_fcf, ev_ntm, sbc):
    """Calculate Martin Stock Score."""
    k = growth if growth else 0  # Growth %
    q = nrr if nrr else 1.0      # NRR (as decimal, e.g. 1.2 for 120%)
    o = rule40_fcf if rule40_fcf else 0  # Rule 40 FCF %
    p = ev_ntm if ev_ntm else 0  # EV/NTM
    r = sbc if sbc else 0        # SBC %
    
    # Engine: Growth + (NRR - 100), capped at 35
    engine = min(35, (k * 100) + ((q * 100) - 100))
    
    # Fuel: Rule 40 / 2, capped at 30
    fuel = min(30, (o * 100) / 2)
    
    # Price Score
    price_score = 0
    if p < 6: price_score = 20
    elif p < 10: price_score = 12
    elif p < 15: price_score = 6
    
    # Discipline: SBC / 2, capped at 15
    discipline = max(0, 15 - ((r * 100) / 2))
    
    return round(engine + fuel + price_score + discipline, 2)

def get_rating(mss):
    if mss >= 80: return "STRONG BUY"
    if mss >= 60: return "BUY"
    if mss >= 40: return "HOLD"
    return "SELL"

# --- Data Processing ---

def normalize_percent(val_str):
    """Convert raw Excel value to decimal (0.30 for 30%)."""
    if val_str is None:
        return None
    if isinstance(val_str, float):
        # Excel stores 0.235 for 23.5% typically, but sometimes 23.5 for 23.5%
        if abs(val_str) <= 1.0 and val_str > -1.0:
            return val_str
        return val_str / 100.0
    return None

def load_excel_data():
    """Load data from Excel SaaS sheet."""
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl not installed. Run: pip install openpyxl")
        return [], []
    
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    if "SaaS" not in wb.sheetnames:
        print(f"Error: 'SaaS' sheet not found. Available: {wb.sheetnames}")
        return [], []
    
    ws = wb["SaaS"]
    headers = [cell.value for cell in ws[1]]
    
    # Filter out None headers
    valid_headers = [h for h in headers if h is not None]
    
    data = []
    for row in ws.iter_rows(min_row=2):
        row_data = {}
        for i, cell in enumerate(row):
            if headers[i] is not None:
                row_data[headers[i]] = cell.value
        data.append(row_data)
    
    return valid_headers, data

def enrich_and_sort_data(headers, data):
    """Calculate missing formulas, enrich data, and sort."""
    
    # Find column indices/names (filter out None headers)
    valid_headers = [h for h in headers if h is not None]
    col = {h.lower().replace('\n', ' ').strip(): h for h in valid_headers}
    
    growth_col = col.get('q. rev growth (yoy)')
    ops_col = col.get('ops margin')
    fcf_col = col.get('fcf margin')
    sbc_col = col.get('sbc % rev')
    ps_col = col.get('p/s (ttm)')
    rule40_fcf_col = col.get('rule of 40 (fcf)')
    ev_ntm_col = col.get('ev / ntm rev')
    nrr_col = col.get('nrr')
    
    enriched = []
    for row in data:
        new_row = row.copy()
        
        # Normalize Quarter format
        if 'Quarter' in new_row:
            new_row['Quarter'] = normalize_quarter(new_row.get('Quarter'))
        
        # Parse values
        growth = normalize_percent(row.get(growth_col))
        ops = normalize_percent(row.get(ops_col))
        fcf = normalize_percent(row.get(fcf_col))
        sbc = normalize_percent(row.get(sbc_col))
        ps = normalize_percent(row.get(ps_col))
        ev_ntm = normalize_percent(row.get(ev_ntm_col))
        nrr = normalize_percent(row.get(nrr_col))
        rule40_fcf_existing = normalize_percent(row.get(rule40_fcf_col))
        
        # Calculate Rule of 40
        rule40 = calculate_rule_of_40(growth, ops)
        if rule40 is not None:
            new_row['Rule of 40'] = round(rule40, 4)
        
        # Calculate Rule of 40 (FCF) if missing
        rule40_fcf_calc = rule40_fcf_existing or calculate_rule_of_40_fcf(growth, fcf)
        if rule40_fcf_calc is not None:
            new_row['Rule of 40 (FCF)'] = round(rule40_fcf_calc, 4)
        
        # Calculate Dilution-Adj FCF Margin
        dil_adj_fcf = calculate_dilution_adj_fcf(fcf, sbc)
        if dil_adj_fcf is not None:
            new_row['Dilution-Adj FCF Margin'] = round(dil_adj_fcf, 4)
        
        # Calculate P/S to Rof40
        ps_to_rof40 = calculate_ps_to_rof40(ps, rule40)
        if ps_to_rof40 is not None:
            new_row['P/S to Rof40'] = round(ps_to_rof40, 2)
        
        # Calculate P/S to Growth
        ps_to_growth = calculate_ps_to_growth(ps, growth)
        if ps_to_growth is not None:
            new_row['P/S to Growth'] = round(ps_to_growth, 2)
        
        # Calculate MSS Score
        mss = calculate_mss(growth, nrr, rule40_fcf_calc, ev_ntm, sbc)
        new_row['MSS Score'] = mss
        
        # Calculate MSS Rating
        new_row['MSS Rating'] = get_rating(mss)
        
        enriched.append(new_row)
    
    # Filter out rows without Stock/Ticker
    enriched = [row for row in enriched if row.get('Stock') or row.get('Ticker')]
    
    # Sort by Ticker A-Z, then Quarter old->new
    enriched.sort(key=lambda x: (
        (x.get('Stock') or x.get('Ticker') or '').upper(),
        get_quarter_order(x.get('Quarter'))
    ))
    
    return enriched

def generate_csv(enriched_data):
    """Generate CSV backup."""
    if not enriched_data:
        return
    headers = list(enriched_data[0].keys())
    with open(CSV_FILE, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(enriched_data)
    print(f"CSV backup: {CSV_FILE}")

def generate_markdown(enriched_data):
    """Generate Obsidian Markdown table."""
    data = [row for row in enriched_data if row.get('Stock') or row.get('Ticker')]
    
    data.sort(key=lambda x: (
        (x.get('Stock') or x.get('Ticker') or '').upper(),
        get_quarter_order(x.get('Quarter') or x.get('Q3/2025'))
    ))
    
    if not data:
        print("No data to generate!")
        return
    
    headers = list(data[0].keys())
    
    md = f"# SaaS Equity Dashboard\n\n"
    md += f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M')} | **Source:** MK Stock Analysis.xlsx (SaaS Sheet)\n\n"
    md += f"**Total Stocks:** {len(data)}\n\n"
    
    md += "| " + " | ".join(str(h) for h in headers) + " |\n"
    md += "| " + " | ".join(["---"] * len(headers)) + " |\n"
    
    for row in data:
        values = []
        for h in headers:
            val = row.get(h)
            header_lower = (h or "").lower().replace('\n', ' ')
            if any(kw in header_lower for kw in ['margin', 'growth', 'nrr', 'sbc', 'revenue', 'score']):
                values.append(format_percent(val))
            elif any(kw in header_lower for kw in ['price', 'value', 'target', 'ev', 'market', 'dma']):
                values.append(format_currency(val))
            else:
                values.append(clean_value(val))
        md += "| " + " | ".join(str(v) for v in values) + " |\n"
    
    with open(MARKDOWN_FILE, 'w') as f:
        f.write(md)
    print(f"Markdown: {MARKDOWN_FILE}")

def main():
    print("Loading Excel data...")
    headers, data = load_excel_data()
    
    if not headers:
        return
    
    print(f"Found {len(data)} rows in SaaS sheet.")
    print("Calculating formulas...")
    enriched_data = enrich_and_sort_data(headers, data)
    print("Generating CSV...")
    generate_csv(enriched_data)
    print("Generating Markdown...")
    generate_markdown(enriched_data)
    print("Done!")

if __name__ == "__main__":
    main()
